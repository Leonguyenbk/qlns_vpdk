"""Xuất danh sách nhân sự ra Excel theo đúng bố cục biểu "Phụ lục 4".

build_workbook(employees, units_by_id, ...) -> BytesIO (nội dung file .xlsx)

- Nhóm theo cây đơn vị: Văn phòng tỉnh (Ban Giám đốc → các phòng → các tổ) rồi
  từng chi nhánh (Ban Giám đốc → các bộ phận).
- Dòng tiêu đề đơn vị: cột A = số thứ tự nhóm/mục, cột B = tên, cột J = tổng số người.
- Người: đánh số 1..N liên tục trong mỗi phòng / bộ phận (tổ chỉ là dòng chia nhỏ).
- Danh sách phụ thuộc bộ lọc truyền vào (đã lọc ở tầng service); không lọc = tất cả.
"""
from __future__ import annotations

import io
from collections import OrderedDict
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_FONT_NAME = "Times New Roman"


def _font(size: int = 12, **kw) -> Font:
    return Font(name=_FONT_NAME, size=size, **kw)

# A..S = 19 cột
COLS = 19
_HEAD_FILL = PatternFill("solid", fgColor="E8EEF7")
_GROUP_FILL = PatternFill("solid", fgColor="DDE7F3")
_SECTION_FILL = PatternFill("solid", fgColor="F1F5FB")
_thin = Side(style="thin", color="B7C3D6")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X",
          "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX",
          "XXI", "XXII", "XXIII", "XXIV", "XXV", "XXVI", "XXVII", "XXVIII", "XXIX", "XXX"]


def _d(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if not v:
        return ""
    return str(v)[:10]


def _chain_to_group(unit, by_id):
    """[nhóm, cấp 2?, cấp 3?] — đi ngược tới HEAD_OFFICE/BRANCH gần nhất."""
    chain = []
    node = unit
    seen = set()
    while node is not None and node.id not in seen:
        seen.add(node.id)
        chain.append(node)
        if node.unit_type in ("HEAD_OFFICE", "BRANCH"):
            break
        node = by_id.get(node.parent_id)
    chain.reverse()
    return chain


def _si(u):
    return u.sort_index if u is not None and u.sort_index is not None else 10 ** 9


def build_workbook(
    employees,
    units_by_id,
    *,
    as_of: date | None = None,
    include_identity: bool = False,
    filter_note: str | None = None,
    org_line1: str = "SỞ NÔNG NGHIỆP VÀ MÔI TRƯỜNG",
    org_line2: str = "VĂN PHÒNG ĐĂNG KÝ ĐẤT ĐAI",
) -> io.BytesIO:
    as_of = as_of or date.today()

    # ---- Gom cây: nhóm -> mục -> (tổ) ----
    ROOT = "__root__"
    groups: "OrderedDict" = OrderedDict()

    for emp in employees:
        a = emp.primary_active_assignment()
        unit = a.unit if a else None
        chain = _chain_to_group(unit, units_by_id) if unit else []
        grp = chain[0] if chain else None
        sec = chain[1] if len(chain) >= 2 else None
        team = chain[2] if len(chain) >= 3 else None

        g_key = grp.id if grp else 0
        gnode = groups.get(g_key)
        if gnode is None:
            gnode = groups[g_key] = {"unit": grp, "sections": OrderedDict()}
        s_key = sec.id if sec else ROOT
        snode = gnode["sections"].get(s_key)
        if snode is None:
            snode = gnode["sections"][s_key] = {"unit": sec, "direct": [], "teams": OrderedDict()}
        if team is not None:
            tnode = snode["teams"].get(team.id)
            if tnode is None:
                tnode = snode["teams"][team.id] = {"unit": team, "people": []}
            tnode["people"].append((emp, a))
        else:
            snode["direct"].append((emp, a))

    # sắp lại theo sort_index cho chắc
    groups = OrderedDict(sorted(groups.items(), key=lambda kv: _si(kv[1]["unit"])))
    for gnode in groups.values():
        gnode["sections"] = OrderedDict(
            sorted(gnode["sections"].items(), key=lambda kv: _si(kv[1]["unit"]))
        )
        for snode in gnode["sections"].values():
            snode["teams"] = OrderedDict(
                sorted(snode["teams"].items(), key=lambda kv: _si(kv[1]["unit"]))
            )

    def _sec_count(snode):
        return len(snode["direct"]) + sum(len(t["people"]) for t in snode["teams"].values())

    def _grp_count(gnode):
        return sum(_sec_count(s) for s in gnode["sections"].values())

    grand_total = sum(_grp_count(g) for g in groups.values())

    # ---- Vẽ workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Phu luc 4"
    ws.sheet_view.showGridLines = False

    widths = [5, 24, 8, 8, 26, 26, 15, 12, 18, 20, 24, 16, 14, 16, 12, 20, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_format.defaultRowHeight = 18

    def put(r, c, val, *, font=None, align=None, fill=None, border=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font or _font(size=12)
        if align:
            cell.alignment = align
        if fill:
            cell.fill = fill
        if border:
            cell.border = _BORDER
        return cell

    def span(r1, c1, r2, c2, val, **kw):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        put(r1, c1, val, **kw)
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = _BORDER

    b = _font(bold=True)
    b12 = _font(bold=True, size=13)

    # Tiêu đề
    ws.merge_cells("A1:E1"); put(1, 1, org_line1, font=b, align=_LEFT, border=False)
    ws.merge_cells("A2:E2"); put(2, 1, org_line2, font=b, align=_LEFT, border=False)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=COLS)
    put(4, 1, "Phụ lục 4", align=Alignment(horizontal="right"), font=_font(italic=True), border=False)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=COLS)
    put(5, 1, "THỐNG KÊ VIÊN CHỨC VÀ NGƯỜI LAO ĐỘNG VĂN PHÒNG ĐĂNG KÝ ĐẤT ĐAI",
        align=_CENTER, font=b12, border=False)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=COLS)
    put(6, 1, f"(Số liệu thống kê cập nhật tính đến ngày {as_of.strftime('%d/%m/%Y')})",
        align=_CENTER, font=_font(italic=True), border=False)
    if filter_note:
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=COLS)
        put(7, 1, f"Danh sách đã lọc: {filter_note}", align=_CENTER,
            font=_font(italic=True, color="1F4E9C"), border=False)

    # Tiêu đề cột (dòng 8-10)
    HR1, HR3 = 8, 10
    hdr = _font(bold=True, size=11)
    simple = {
        1: "TT", 2: "Họ và tên", 5: "Quê quán (số nhà, TDP; thôn; xã, phường; huyện, tỉnh)",
        6: "Nơi ở hiện nay", 7: "Số CCCD", 8: "Ngày cấp", 9: "Nơi cấp",
        10: "Chức vụ, chức danh", 11: "Các nhiệm vụ đang đảm nhận",
        12: "Ngày vào biên chế (tuyển dụng, HĐLĐ)", 13: "Loại Hợp đồng",
        14: "Ngạch (Chức danh nghề nghiệp) hiện đang giữ",
    }
    for c, t in simple.items():
        span(HR1, c, HR3, c, t, font=hdr, align=_CENTER, fill=_HEAD_FILL)
    span(HR1, 3, HR1, 4, "Năm sinh", font=hdr, align=_CENTER, fill=_HEAD_FILL)
    span(HR1 + 1, 3, HR3, 3, "Nam", font=hdr, align=_CENTER, fill=_HEAD_FILL)
    span(HR1 + 1, 4, HR3, 4, "Nữ", font=hdr, align=_CENTER, fill=_HEAD_FILL)
    span(HR1, 15, HR1 + 1, 17, "Trình độ chuyên môn cao nhất", font=hdr, align=_CENTER, fill=_HEAD_FILL)
    span(HR1, 18, HR1 + 1, 19, "Chứng chỉ", font=hdr, align=_CENTER, fill=_HEAD_FILL)
    for c, t in {15: "Trình độ", 16: "Ngành đào tạo", 17: "Hệ đào tạo",
                 18: "Ngoại ngữ", 19: "Tin học"}.items():
        put(HR3, c, t, font=hdr, align=_CENTER, fill=_HEAD_FILL)
    # dòng số thứ tự cột
    r = HR3 + 1
    for c in range(1, COLS + 1):
        put(r, c, c, font=_font(italic=True, size=9), align=_CENTER, fill=_HEAD_FILL)
    ws.freeze_panes = ws.cell(row=r + 1, column=1)

    # ---- Dữ liệu ----
    r += 1

    def person_row(rr, tt, emp, a):
        pos_title = emp.professional_title or (a.position.name if a and a.position else "")
        vals = {
            1: tt, 2: emp.full_name,
            3: _d(emp.date_of_birth) if emp.gender == "MALE" else "",
            4: _d(emp.date_of_birth) if emp.gender == "FEMALE" else "",
            5: emp.place_of_origin or "", 6: emp.address or "",
            7: (emp.identity_number or "") if include_identity else "",
            8: _d(emp.identity_issued_date), 9: emp.identity_issued_place or "",
            10: pos_title, 11: emp.job_duties or "",
            12: _d(emp.tenure_date) + (
                f" ({_d(emp.recruitment_date)})"
                if emp.recruitment_date and emp.recruitment_date != emp.tenure_date else ""
            ),
            13: emp.contract_type or "", 14: emp.job_grade_code or "",
            15: emp.education_level or "", 16: emp.education_major or "",
            17: emp.education_mode or "", 18: emp.foreign_language_cert or "",
            19: emp.it_cert or "",
        }
        for c in range(1, COLS + 1):
            put(rr, c, vals.get(c, ""),
                align=_CENTER if c in (1, 3, 4, 8, 12) else _LEFT_TOP, font=_font(size=12))

    def header_row(rr, label_a, name, count, fill, bold=True):
        put(rr, 1, label_a, font=_font(bold=bold), align=_CENTER, fill=fill)
        put(rr, 2, name, font=_font(bold=bold), align=_LEFT, fill=fill)
        for c in range(3, COLS + 1):
            put(rr, c, "", fill=fill)
        put(rr, 10, count if count else "", font=_font(bold=bold), align=_CENTER, fill=fill)

    gletter = ord("A")
    for gnode in groups.values():
        gname = gnode["unit"].name if gnode["unit"] else "Chưa phân công"
        header_row(r, chr(gletter), gname, _grp_count(gnode), _GROUP_FILL)
        gletter += 1
        r += 1
        ri = 0
        for skey, snode in gnode["sections"].items():
            sname = snode["unit"].name if snode["unit"] else "Trực thuộc"
            roman = _ROMAN[ri] if ri < len(_ROMAN) else str(ri + 1)
            ri += 1
            header_row(r, roman, sname, _sec_count(snode), _SECTION_FILL)
            r += 1
            n = 0
            # người trực thuộc mục (không thuộc tổ nào)
            if snode["direct"]:
                if snode["teams"]:
                    header_row(r, "-", "(Trực thuộc)", len(snode["direct"]),
                               PatternFill("solid", fgColor="FAFCFF"), bold=False)
                    r += 1
                for emp, a in snode["direct"]:
                    n += 1
                    person_row(r, n, emp, a)
                    r += 1
            # từng tổ
            for tnode in snode["teams"].values():
                header_row(r, "-", tnode["unit"].name, len(tnode["people"]),
                           PatternFill("solid", fgColor="FAFCFF"), bold=False)
                r += 1
                for emp, a in tnode["people"]:
                    n += 1
                    person_row(r, n, emp, a)
                    r += 1

    # Tổng cộng
    header_row(r, "", "Tổng cộng", grand_total, _GROUP_FILL)
    r += 2

    put(r, 13, f"……, ngày …… tháng …… năm {as_of.year}",
        align=Alignment(horizontal="center"), font=_font(italic=True), border=False)
    r += 1
    put(r, 3, "NGƯỜI LẬP", align=Alignment(horizontal="center"), font=b, border=False)
    put(r, 13, "THỦ TRƯỞNG ĐƠN VỊ", align=Alignment(horizontal="center"), font=b, border=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
