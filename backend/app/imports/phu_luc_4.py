"""Nhập danh sách viên chức / người lao động từ biểu mẫu "Phụ lục 4".

Biểu mẫu là 1 sheet Excel, cấu trúc phân cấp bằng cột "TT":
  - chữ cái (A, B, ...)  -> nhóm cấp cao (Văn phòng tỉnh / Khối chi nhánh)
  - số La Mã (I, II, ...) -> đơn vị / bộ phận
  - số (1, 2, ...)        -> một người
Ô cột 10 (J) trên dòng tiêu đề đơn vị chứa TỔNG số người của đơn vị đó -> dùng để đối soát.

Cách dùng:
    from app.imports.phu_luc_4 import import_workbook
    report = import_workbook("Phụ lục 4.xlsx", dry_run=True)
    print(report.render())

Hoặc CLI:  flask --app wsgi import-phuluc4 "Phụ lục 4.xlsx" --dry-run
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import load_workbook

from ..extensions import db
from ..models import Employee, EmployeeAssignment, EmployeeEducation, JobGrade, OrganizationUnit, Position

# ── Vị trí cột (0-based) trong sheet ────────────────────────────────────────────
C_TT, C_NAME = 0, 1
C_BY_MALE, C_BY_FEMALE = 2, 3
C_ORIGIN, C_RESIDENCE = 4, 5
C_CCCD, C_CCCD_DATE, C_CCCD_PLACE = 6, 7, 8
C_TITLE, C_DUTIES = 9, 10
C_JOIN_DATE, C_CONTRACT, C_GRADE = 11, 12, 13
C_EDU_LEVEL, C_EDU_MAJOR, C_EDU_MODE = 14, 15, 16
C_LANG, C_IT = 17, 18

# So khớp với chuỗi đã _strip_accents().lower() -> viết không dấu.
FOOTER_MARKERS = ("tong cong", "nguoi lap", "thu truong don vi", "thang      nam")

# Loại hợp đồng (giá trị gốc) -> (nhãn chuẩn, employment_type enum hiện có)
CONTRACT_MAP = {
    "viên chức": ("Viên chức", "OFFICIAL"),
    "kxđth": ("HĐLĐ không xác định thời hạn", "CONTRACT"),
    "không xác định thời hạn": ("HĐLĐ không xác định thời hạn", "CONTRACT"),
    "12 tháng": ("HĐLĐ 12 tháng", "CONTRACT"),
    "thử việc 2 tháng": ("Thử việc 2 tháng", "PROBATION"),
    "thử việc": ("Thử việc", "PROBATION"),
}

EDU_RANK = {"ts": 5, "tskh": 6, "ths": 4, "th.s": 4, "đh": 3, "cđ": 2, "tc": 1, "sc": 0}

# Chức vụ (chuỗi tự do) -> (tên chuẩn, is_managerial, level). level nhỏ = chức vụ cao.
TITLE_RULES = [
    ("phó giám đốc", "Phó Giám đốc", True, 20),
    ("giám đốc", "Giám đốc", True, 10),
    ("phó trưởng phòng", "Phó Trưởng phòng", True, 35),
    ("trưởng phòng", "Trưởng phòng", True, 30),
    ("phó trưởng bộ phận", "Phó Trưởng bộ phận", True, 45),
    ("trưởng bộ phận", "Trưởng bộ phận", True, 40),
    ("kế toán trưởng", "Kế toán trưởng", True, 38),
    ("tổ trưởng", "Tổ trưởng", True, 50),
    ("tổ phó", "Tổ phó", True, 55),
    ("thủ quỹ", "Thủ quỹ", False, 85),
    ("văn thư", "Văn thư", False, 85),
    ("chuyên viên", "Chuyên viên", False, 80),
    ("nhân viên", "Nhân viên", False, 90),
    ("viên chức", "Chuyên viên", False, 80),
]
_DEFAULT_TITLE = ("Nhân viên", False, 90)

# Chia thêm cấp "Tổ" bên trong các phòng của Văn phòng tỉnh.
# key = tên phòng; value = [(từ khoá không dấu trong "nhiệm vụ"/"chức vụ", tên Tổ chuẩn)]
# Khớp theo thứ tự; không khớp -> người ở thẳng phòng. Lãnh đạo phòng luôn ở thẳng phòng.
TEAM_RULES = {
    "Phòng Tổ chức - Hành chính": [
        ("to chuc bo may", "Tổ Tổ chức bộ máy, nhân sự"),
        ("quan tri hanh chinh", "Tổ Quản trị hành chính - văn phòng"),
        ("van thu", "Tổ Quản trị hành chính - văn phòng"),
        ("phap che", "Tổ Pháp chế"),
        ("kiem soat nghiep vu", "Tổ Kiểm soát nghiệp vụ"),
    ],
    "Phòng Kế hoạch - Tài chính": [
        ("ke hoach", "Tổ Kế hoạch"),
        ("tai chinh", "Tổ Tài chính"),
        ("ke toan", "Tổ Tài chính"),
    ],
    "Phòng Kỹ thuật - Đăng ký đất đai": [
        ("giai quyet thu tuc hanh chinh", "Tổ Giải quyết thủ tục hành chính"),
        ("giai quyet tthc", "Tổ Giải quyết thủ tục hành chính"),
        ("kiem tra, giam sat", "Tổ Kiểm tra, giám sát"),
        ("kiem tra giam sat", "Tổ Kiểm tra, giám sát"),
        ("ban do", "Tổ Bản đồ"),
    ],
    "Phòng Dữ liệu - thông tin đất đai": [
        ("ung dung va phat trien cong nghe", "Tổ Ứng dụng và Phát triển công nghệ"),
        ("co so du lieu", "Tổ Cơ sở dữ liệu"),
        ("thong tin luu tru", "Tổ Thông tin - Lưu trữ"),
        ("thong tin - luu tru", "Tổ Thông tin - Lưu trữ"),
    ],
}

_INT = re.compile(r"^\d+$")
_DATE_TOKEN = re.compile(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})|(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})")
_YEAR_ONLY = re.compile(r"(?:năm\s*)?((?:19|20)\d{2})", re.IGNORECASE)


def _strip_accents(s: str) -> str:
    # đ/Đ không bị NFD tách -> thay thủ công trước
    s = s.replace("đ", "d").replace("Đ", "D")
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _slug(s: str, maxlen: int = 40) -> str:
    s = _strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:maxlen] or "x"


def _norm(v) -> str | None:
    if v is None:
        return None
    s = str(v).replace("\n", " ").replace("\xa0", " ").strip().strip("/").strip()
    s = re.sub(r"\s{2,}", " ", s)
    return s or None


def _clean_cccd(v) -> str | None:
    s = _norm(v)
    if not s:
        return None
    return re.sub(r"[^0-9A-Za-z]", "", s) or None


def _parse_dates(v) -> list[date]:
    """Trả về danh sách ngày tìm thấy trong ô (ô có thể chứa 2 ngày)."""
    if isinstance(v, datetime):
        return [v.date()]
    if isinstance(v, date):
        return [v]
    s = _norm(v)
    if not s:
        return []
    out: list[date] = []
    for m in _DATE_TOKEN.finditer(s):
        try:
            if m.group(1):  # yyyy-mm-dd
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            else:           # dd/mm/yyyy — nếu tháng > 12 thì file ghi kiểu mm/dd/yyyy
                d, mo, y = int(m.group(4)), int(m.group(5)), int(m.group(6))
                if mo > 12 and d <= 12:
                    d, mo = mo, d
                if y < 100:
                    y += 1900 if y > 30 else 2000
            out.append(date(y, mo, d))
        except (ValueError, TypeError):
            continue
    if not out:  # chỉ có năm ("Năm 2016")
        ym = _YEAR_ONLY.search(s)
        if ym:
            out.append(date(int(ym.group(1)), 1, 1))
    return out


def _split_multi(v) -> list[str]:
    s = _norm(v)
    if not s:
        return []
    return [p.strip() for p in re.split(r"[,;/]|\band\b", s) if p.strip()]


@dataclass
class ImportReport:
    dry_run: bool = True
    source: str = ""
    total_people_rows: int = 0
    employees_created: int = 0
    employees_updated: int = 0
    units_created: int = 0
    positions_created: int = 0
    grades_seen: set = field(default_factory=set)
    unparsed_dates: list = field(default_factory=list)
    headcount_mismatch: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    def render(self) -> str:
        lines = [
            f"Nguồn: {self.source}   ({'DRY-RUN — không ghi DB' if self.dry_run else 'ĐÃ GHI DB'})",
            f"  Dòng người đọc được : {self.total_people_rows}",
            f"  Nhân sự tạo mới      : {self.employees_created}",
            f"  Nhân sự cập nhật     : {self.employees_updated}",
            f"  Đơn vị tạo mới       : {self.units_created}",
            f"  Chức vụ tạo mới      : {self.positions_created}",
            f"  Mã ngạch gặp phải    : {len(self.grades_seen)}  ({', '.join(sorted(self.grades_seen))})",
        ]
        if self.headcount_mismatch:
            lines.append(f"  ⚠ Lệch số lượng ({len(self.headcount_mismatch)}):")
            lines += [f"      {u}: sheet={exp}, đọc được={got}" for u, exp, got in self.headcount_mismatch]
        if self.unparsed_dates:
            lines.append(f"  ⚠ Ngày không đọc được ({len(self.unparsed_dates)}):")
            lines += [f"      dòng {r} [{col}]: {val!r}" for r, col, val in self.unparsed_dates[:40]]
        if self.warnings:
            lines.append(f"  ⚠ Cảnh báo khác ({len(self.warnings)}):")
            lines += [f"      {w}" for w in self.warnings[:40]]
        return "\n".join(lines)


class _Importer:
    def __init__(self, path: str, sheet: str | None, as_of: date, report: ImportReport):
        self.wb = load_workbook(path, data_only=True, read_only=True)
        self.ws = self.wb[sheet] if sheet else self.wb.active
        self.as_of = as_of
        self.r = report
        self._unit_cache: dict[tuple[int | None, str], OrganizationUnit] = {}
        self._pos_cache: dict[str, Position] = {}
        self._next_code = self._max_employee_seq()

    # ---- helpers ------------------------------------------------------------
    def _max_employee_seq(self) -> int:
        mx = 0
        for (code,) in db.session.query(Employee.employee_code).all():
            m = re.match(r"^NS(\d+)$", code or "")
            if m:
                mx = max(mx, int(m.group(1)))
        return mx

    def _gen_code(self) -> str:
        self._next_code += 1
        return f"NS{self._next_code:04d}"

    def get_unit(self, name: str, unit_type: str, parent: OrganizationUnit | None) -> OrganizationUnit:
        name = re.sub(r"\bBô phận\b", "Bộ phận", name).strip()
        key = (parent.id if parent else None, name.lower())
        if key in self._unit_cache:
            return self._unit_cache[key]
        pid = parent.id if parent else None
        unit = (
            db.session.query(OrganizationUnit)
            .filter(OrganizationUnit.name == name, OrganizationUnit.parent_id.is_(pid) if pid is None
                    else OrganizationUnit.parent_id == pid)
            .first()
        )
        if unit is None:
            base = _slug(name)
            code = (f"{parent.code}." if parent else "") + base
            code = code[:50]
            n = 1
            while db.session.query(OrganizationUnit).filter_by(code=code).first():
                n += 1
                code = (code[:46] + f"-{n}")[:50]
            unit = OrganizationUnit(code=code, name=name, unit_type=unit_type, parent_id=pid)
            db.session.add(unit)
            db.session.flush()
            self.r.units_created += 1
        self._unit_cache[key] = unit
        return unit

    def get_position(self, title_raw: str | None) -> tuple[Position, bool]:
        t = (title_raw or "").lower()
        canon, managerial, level = _DEFAULT_TITLE
        for needle, name, mgr, lv in TITLE_RULES:
            if needle in t:
                canon, managerial, level = name, mgr, lv
                break
        if canon in self._pos_cache:
            return self._pos_cache[canon], managerial
        pos = db.session.query(Position).filter_by(name=canon).first()
        if pos is None:
            code = _slug(canon, 30)
            n = 1
            while db.session.query(Position).filter_by(code=code).first():
                n += 1
                code = f"{_slug(canon, 26)}-{n}"
            pos = Position(code=code, name=canon, is_managerial=managerial, level=level)
            db.session.add(pos)
            db.session.flush()
            self.r.positions_created += 1
        else:
            # đồng bộ lại thứ hạng cho chức vụ đã có (an toàn: chỉ các tên chuẩn)
            if pos.level != level:
                pos.level = level
            pos.is_managerial = managerial
        self._pos_cache[canon] = pos
        return pos, managerial

    def resolve_team(self, unit: OrganizationUnit, title_raw: str | None, duties_raw: str | None) -> OrganizationUnit:
        """Nếu người thuộc 1 phòng của Văn phòng tỉnh -> tách xuống Tổ theo nhiệm vụ."""
        rules = TEAM_RULES.get(unit.name)
        if not rules:
            return unit
        text = _strip_accents(f"{duties_raw or ''} | {title_raw or ''}").lower()
        if "truong phong" in text or "pho truong phong" in text:
            return unit  # lãnh đạo phòng ở thẳng phòng
        for needle, team_name in rules:
            if needle in text:
                return self.get_unit(team_name, "SECTION", unit)
        return unit

    def touch_grade(self, code: str | None):
        if not code:
            return
        self.r.grades_seen.add(code)
        if not db.session.query(JobGrade).filter_by(code=code).first():
            cat = "Hợp đồng" if re.fullmatch(r"[IVX]+", code) else "Viên chức"
            db.session.add(JobGrade(code=code, category=cat))

    # ---- phân loại 1 dòng: người / đơn vị / nhãn -----------------------------
    @staticmethod
    def _is_person_row(tt: str | None, name: str | None, row) -> bool:
        if not (tt and _INT.match(tt)):
            return False
        if not name or _INT.match(name):
            return False  # dòng đánh số cột "1 | 2 | 3 ..."
        def cell(i):
            return row[i] if len(row) > i else None
        # người thật: có CCCD, hoặc có năm sinh, hoặc có chức vụ/chức danh
        return bool(_clean_cccd(cell(C_CCCD)) or _norm(cell(C_BY_MALE)) or _norm(cell(C_BY_FEMALE))
                    or _norm(cell(C_TITLE)))

    def _classify_unit(self, name: str, head, branch):
        """Trả về (unit_hoặc_None, new_branch, new_section). Phân loại theo TÊN, không theo số thứ tự."""
        nm = re.sub(r"\bBô phận\b", "Bộ phận", name).strip()
        low = _strip_accents(nm).lower()
        if low.startswith("van phong dang ky"):
            h = self.get_unit(nm, "HEAD_OFFICE", None)
            return h, None, None                       # head; xoá branch & section
        if low.startswith("cn ") or low.startswith("chi nhanh"):
            b = self.get_unit(nm, "BRANCH", head)
            return None, b, None                        # branch mới; section = None
        if low.startswith("phong "):
            s = self.get_unit(nm, "DEPARTMENT", head)
            return s, None, s                           # phòng thuộc tỉnh -> quay lại ngữ cảnh head
        if low.startswith("lanh dao chi nhanh") or (branch is not None and low.startswith("ban giam doc")):
            # Ban giám đốc chi nhánh — ngang hàng các bộ phận của chi nhánh
            s = self.get_unit("Ban Giám đốc", "SECTION", branch or head)
            return s, branch, s
        if low.startswith("lanh dao van phong") or low.startswith("ban giam doc"):
            # Ban Giám đốc (văn phòng tỉnh) — ngang hàng các phòng
            s = self.get_unit("Ban Giám đốc", "DEPARTMENT", head)
            return s, None, s
        if low.startswith("bo phan ") or low.startswith("to "):
            s = self.get_unit(nm, "SECTION", branch or head)
            return s, branch, s
        if low in ("khoi chi nhanh", "cac chi nhanh", "chi nhanh"):
            return None, None, None                     # nhãn phân nhóm, bỏ qua
        # không rõ -> coi là bộ phận thuộc ngữ cảnh hiện tại
        s = self.get_unit(nm, "SECTION", branch or head)
        return s, branch, s

    # ---- main pass --------------------------------------------------------
    def run(self):
        head = branch = section = None
        expected: dict[int, int] = {}   # unit_id -> headcount ghi trong sheet
        actual: dict[int, int] = {}     # unit_id -> số người thực đọc
        unit_name: dict[int, str] = {}

        for idx, row in enumerate(self.ws.iter_rows(values_only=True), start=1):
            tt = _norm(row[C_TT] if len(row) > C_TT else None)
            name = _norm(row[C_NAME] if len(row) > C_NAME else None)
            if not tt and not name:
                continue
            low_name = _strip_accents(name or "").lower()
            if not tt and any(mk in low_name for mk in FOOTER_MARKERS):
                break
            if (tt or "").upper() == "TT" or low_name in ("ho va ten",):
                continue

            if self._is_person_row(tt, name, row):
                unit = section or branch or head
                if unit is None:
                    self.r.warnings.append(f"dòng {idx}: người '{name}' chưa xác định được đơn vị")
                    continue
                self.r.total_people_rows += 1
                actual[unit.id] = actual.get(unit.id, 0) + 1
                self._upsert_person(idx, row, unit)
                continue

            # dòng còn lại có TÊN ở cột B mà không phải người => tiêu đề đơn vị
            if name and not _INT.match(name):
                u, branch, section = self._classify_unit(name, head, branch)
                if u is not None and u.unit_type == "HEAD_OFFICE":
                    head = u
                # số lượng ghi ở cột 10 (J) trên dòng tiêu đề đơn vị
                hc = row[C_TITLE] if len(row) > C_TITLE else None
                target = section or branch
                if isinstance(hc, (int, float)) and target is not None:
                    expected[target.id] = int(hc)
                    unit_name[target.id] = target.name

        # đối soát số lượng ở mức bộ phận / phòng (bỏ qua cấp tổng hợp)
        for uid, exp in expected.items():
            got = actual.get(uid, 0)
            if got and got != exp:
                self.r.headcount_mismatch.append((unit_name.get(uid, uid), exp, got))
        return self.r

    # ---- person upsert --------------------------------------------------
    def _upsert_person(self, idx: int, row, unit: OrganizationUnit):
        def cell(i):
            return row[i] if len(row) > i else None

        name = _norm(cell(C_NAME))
        gender = "MALE" if _norm(cell(C_BY_MALE)) else ("FEMALE" if _norm(cell(C_BY_FEMALE)) else None)
        dobs = _parse_dates(cell(C_BY_MALE) or cell(C_BY_FEMALE))
        dob = dobs[0] if dobs else None
        if cell(C_BY_MALE) or cell(C_BY_FEMALE):
            if not dob:
                self.r.unparsed_dates.append((idx, "năm sinh", _norm(cell(C_BY_MALE) or cell(C_BY_FEMALE))))

        cccd = _clean_cccd(cell(C_CCCD))
        issued = _parse_dates(cell(C_CCCD_DATE))
        join_dates = _parse_dates(cell(C_JOIN_DATE))
        if cell(C_JOIN_DATE) and not join_dates:
            self.r.unparsed_dates.append((idx, "ngày vào biên chế", _norm(cell(C_JOIN_DATE))))
        tenure_date = join_dates[0] if join_dates else None
        recruitment_date = join_dates[1] if len(join_dates) > 1 else tenure_date

        contract_raw = _norm(cell(C_CONTRACT))
        contract_label, employment_type = None, None
        if contract_raw:
            key = contract_raw.lower()
            for k, (lbl, et) in CONTRACT_MAP.items():
                if k in key:
                    contract_label, employment_type = lbl, et
                    break
            if contract_label is None:
                contract_label = contract_raw

        grade = _norm(cell(C_GRADE))
        self.touch_grade(grade)

        title_raw = _norm(cell(C_TITLE))
        pos, _mgr = self.get_position(title_raw)

        # Tách xuống Tổ nếu người này thuộc một phòng của Văn phòng tỉnh
        unit = self.resolve_team(unit, title_raw, _norm(cell(C_DUTIES)))

        # education
        levels = _split_multi(cell(C_EDU_LEVEL))
        majors = _split_multi(cell(C_EDU_MAJOR))
        mode = _norm(cell(C_EDU_MODE))
        edu_rows = []
        if len(levels) == len(majors) and levels:
            pairs = list(zip(levels, majors))
        elif levels:
            pairs = [(lv, _norm(cell(C_EDU_MAJOR))) for lv in levels]
        else:
            pairs = [(None, _norm(cell(C_EDU_MAJOR)))] if _norm(cell(C_EDU_MAJOR)) else []
        best_rank, best = -1, None
        for lv, mj in pairs:
            edu_rows.append((lv, mj, mode))
            rank = EDU_RANK.get((lv or "").lower().replace(" ", ""), 0)
            if rank >= best_rank:
                best_rank, best = rank, (lv, mj, mode)

        # ---- find or create employee (dedup theo CCCD, fallback tên+ngày sinh) ----
        emp = None
        if cccd:
            emp = db.session.query(Employee).filter_by(identity_number=cccd).first()
        if emp is None:
            q = db.session.query(Employee).filter_by(full_name=name)
            if dob:
                q = q.filter_by(date_of_birth=dob)
            emp = q.first() if (dob or not cccd) else None

        creating = emp is None
        if creating:
            emp = Employee(employee_code=self._gen_code(), full_name=name, status="WORKING")
            db.session.add(emp)

        emp.full_name = name or emp.full_name
        emp.gender = gender or emp.gender
        emp.date_of_birth = dob or emp.date_of_birth
        emp.identity_number = cccd or emp.identity_number
        emp.identity_issued_date = (issued[0] if issued else None) or emp.identity_issued_date
        emp.identity_issued_place = _norm(cell(C_CCCD_PLACE)) or emp.identity_issued_place
        emp.place_of_origin = _norm(cell(C_ORIGIN)) or emp.place_of_origin
        emp.address = _norm(cell(C_RESIDENCE)) or emp.address
        emp.professional_title = title_raw or emp.professional_title
        emp.job_duties = _norm(cell(C_DUTIES)) or emp.job_duties
        emp.tenure_date = tenure_date or emp.tenure_date
        emp.recruitment_date = recruitment_date or emp.recruitment_date
        emp.contract_type = contract_label or emp.contract_type
        emp.employment_type = employment_type or emp.employment_type
        emp.job_grade_code = grade or emp.job_grade_code
        emp.foreign_language_cert = _norm(cell(C_LANG)) or emp.foreign_language_cert
        emp.it_cert = _norm(cell(C_IT)) or emp.it_cert
        if best:
            emp.education_level, emp.education_major, emp.education_mode = best
        db.session.flush()

        # education: xoá & dựng lại
        emp.education.clear()
        db.session.flush()
        for i, (lv, mj, md) in enumerate(edu_rows):
            emp.education.append(
                EmployeeEducation(level=lv, major=mj, mode=md, is_highest=(best is not None and (lv, mj, md) == best))
            )

        # phân công chính đang hiệu lực
        current = emp.primary_active_assignment()
        start = recruitment_date or tenure_date or self.as_of
        if current is None:
            db.session.add(EmployeeAssignment(
                employee_id=emp.id, unit_id=unit.id, position_id=pos.id,
                assignment_type="RECRUITMENT", start_date=start, is_primary=True,
                note=title_raw,
            ))
        elif current.unit_id != unit.id or current.position_id != pos.id:
            current.end_date = self.as_of
            db.session.add(EmployeeAssignment(
                employee_id=emp.id, unit_id=unit.id, position_id=pos.id,
                assignment_type="REASSIGNMENT", start_date=self.as_of, is_primary=True,
                note=title_raw,
            ))

        if creating:
            self.r.employees_created += 1
        else:
            self.r.employees_updated += 1


def import_workbook(path: str, *, sheet: str | None = None, dry_run: bool = True,
                    as_of: date | None = None) -> ImportReport:
    report = ImportReport(dry_run=dry_run, source=path)
    imp = _Importer(path, sheet, as_of or date.today(), report)
    try:
        imp.run()
        from ..services.org_index import reindex_units
        reindex_units()  # đánh lại sort_index cho đơn vị
        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    return report
